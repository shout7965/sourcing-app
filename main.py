import os
import re
import io
import json
import requests
import anthropic
import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore as fb_fs, storage as fb_storage
from datetime import datetime, date
from flask import Flask, send_file, request, jsonify, session
from werkzeug.security import generate_password_hash, check_password_hash
from pydantic import BaseModel
from typing import List, Optional
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'sourcing-dev-key')

NAVER_CLIENT_ID     = os.environ.get('NAVER_CLIENT_ID')
NAVER_CLIENT_SECRET = os.environ.get('NAVER_CLIENT_SECRET')
ANTHROPIC_API_KEY   = os.environ.get('ANTHROPIC_API_KEY')
DAILY_LIMIT         = 25_000

# 사용자 목록: APP_USERS_JSON = {"alice": "pass1", "bob": "pass2"}
try:
    APP_USERS = json.loads(os.environ.get('APP_USERS_JSON', '{}'))
except Exception:
    APP_USERS = {}

claude = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

# ── Firebase 초기화 ────────────────────────────────────────────────────────
FIREBASE_ENABLED = False
db = None
STORAGE_BUCKET = ''

def init_firebase():
    global FIREBASE_ENABLED, db, STORAGE_BUCKET
    sa_json = os.environ.get('FIREBASE_SERVICE_ACCOUNT_JSON')
    if not sa_json:
        print("[Firebase] FIREBASE_SERVICE_ACCOUNT_JSON 미설정 → Firestore 기능 비활성화")
        return
    try:
        sa_dict = json.loads(sa_json)
        cred = credentials.Certificate(sa_dict)
        project_id = sa_dict.get('project_id', '')
        bucket_name = f"{project_id}.appspot.com"
        if not firebase_admin._apps:
            firebase_admin.initialize_app(cred, {'storageBucket': bucket_name})
        db = fb_fs.client()
        FIREBASE_ENABLED = True
        STORAGE_BUCKET = bucket_name
        print(f"[Firebase] 초기화 성공 (Storage: {bucket_name})")
    except Exception as e:
        print(f"[Firebase] 초기화 실패: {e}")

init_firebase()

# ── Firebase 헬퍼 ─────────────────────────────────────────────────────────
def _today() -> str:
    return date.today().isoformat()

def increment_usage(count: int = 1):
    if not FIREBASE_ENABLED or count == 0:
        return
    try:
        db.collection('api_usage').document(_today()).set(
            {'count': fb_fs.Increment(count), 'date': _today()}, merge=True
        )
    except Exception as e:
        print(f"[Firebase] usage 저장 실패: {e}")

def _progress_key(keyword: str, source: str) -> str:
    """Firestore 문서 ID용 안전한 키 생성"""
    safe = re.sub(r'[^a-zA-Z0-9가-힣]', '_', keyword)[:30]
    return f"{safe}__{source}"

def save_page_progress(keyword: str, source: str, page: int):
    if not FIREBASE_ENABLED:
        return
    try:
        key = _progress_key(keyword, source)
        ref = db.collection('search_progress').document(key)
        doc = ref.get()
        completed = doc.to_dict().get('completed_pages', []) if doc.exists else []
        if page not in completed:
            completed.append(page)
            completed.sort()
        ref.set({
            'keyword': keyword,
            'source': source,
            'completed_pages': completed,
            'last_updated': fb_fs.SERVER_TIMESTAMP,
        }, merge=True)
    except Exception as e:
        print(f"[Firebase] progress 저장 실패: {e}")

# ── 공통 유틸 ─────────────────────────────────────────────────────────────
def strip_html(text: str) -> str:
    return re.sub(r'<[^>]+>', '', text)

def parse_item_date(item: dict) -> Optional[date]:
    if 'postdate' in item:
        try:
            return datetime.strptime(item['postdate'], '%Y%m%d').date()
        except Exception:
            return None
    if 'datetime' in item:
        try:
            return datetime.fromisoformat(item['datetime']).date()
        except Exception:
            return None
    return None

def format_date(d: Optional[date]) -> str:
    return d.strftime('%Y-%m-%d') if d else ''

def naver_search(endpoint: str, query: str, display: int, start: int, headers: dict) -> tuple[list, int]:
    try:
        resp = requests.get(
            f"https://openapi.naver.com/v1/search/{endpoint}.json",
            headers=headers,
            params={"query": query, "display": display, "start": start, "sort": "date"},
            timeout=10,
        )
        if resp.status_code != 200:
            return [], 0
        data = resp.json()
        return data.get("items", []), int(data.get("total", 0))
    except requests.RequestException:
        return [], 0

def analyze_image_for_product(image_url: str) -> dict:
    """썸네일 이미지에서 영문 브랜드명/제품명 추출 (Claude Vision)"""
    import base64
    try:
        resp = requests.get(image_url, timeout=5, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        })
        if resp.status_code != 200:
            return {}
        content_type = resp.headers.get('content-type', 'image/jpeg').split(';')[0].strip()
        if content_type not in ('image/jpeg', 'image/png', 'image/gif', 'image/webp'):
            content_type = 'image/jpeg'
        image_data = base64.b64encode(resp.content).decode()
        response = claude.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=150,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": content_type, "data": image_data}},
                    {"type": "text", "text": 'What brand and official English product name/model number is shown in this image? Reply ONLY with JSON: {"brand_name": "...", "product_name_en": "..."}. Use null if not clearly visible.'},
                ],
            }],
        )
        text = response.content[0].text.strip()
        text = re.sub(r'```json?\s*|\s*```', '', text).strip()
        return json.loads(text)
    except Exception as e:
        print(f"[Vision] 분석 실패: {e}")
        return {}


def search_naver_shopping(query: str, headers: dict, brand_name: str = None) -> Optional[dict]:
    if not query.strip():
        return None
    try:
        resp = requests.get(
            "https://openapi.naver.com/v1/search/shop.json",
            headers=headers,
            params={"query": query, "display": 5, "sort": "asc"},
            timeout=10,
        )
        if resp.status_code != 200:
            return None
        items = resp.json().get("items", [])
        if not items:
            return None

        # 브랜드명이 있으면 결과 중 브랜드명이 제목에 포함된 항목만 사용
        selected = None
        if brand_name:
            brand_lower = brand_name.lower()
            for it in items:
                title = strip_html(it.get("title", "")).lower()
                if brand_lower in title:
                    selected = it
                    break
            if selected is None:
                return None  # 브랜드 불일치 → 엉뚱한 제품 반환 방지
        else:
            selected = items[0]

        return {
            "shopping_price": int(selected.get("lprice", 0)),
            "shopping_image": strip_html(selected.get("image", "")),
            "shopping_link":  selected.get("link", ""),
            "shopping_mall":  strip_html(selected.get("mallName", "")),
            "shopping_title": strip_html(selected.get("title", "")),
        }
    except Exception:
        return None

# ── Pydantic 스키마 ───────────────────────────────────────────────────────
class ReviewItem(BaseModel):
    index: int
    brand_name:                Optional[str] = None
    product_name:              Optional[str] = None
    product_name_en:           Optional[str] = None   # 영문 공식 제품명/모델번호
    category:                  Optional[str] = None
    purchase_source:           Optional[str] = None
    price_paid:                Optional[str] = None   # 후기에서 언급된 구매 가격
    is_direct_purchase_review: Optional[bool] = None  # 실제 해외직구 후기 여부

class ExtractionResult(BaseModel):
    items: List[ReviewItem]

class SpecItem(BaseModel):
    spec:          str                 # 예: "200g 1개", "200g 2팩", "500g 1봉"
    price:         int                 # 최저가 (원)
    free_shipping: Optional[bool] = None  # True=무료배송 확인, False=유료, None=알수없음
    delivery_fee:  Optional[int]  = None  # 배송비 금액 (원, 확인된 경우)
    seller:        Optional[str]  = None
    link:          Optional[str]  = None

class SpecResult(BaseModel):
    specs: List[SpecItem]

# ── 라우트 ────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return send_file('src/index.html')

# 로그인
@app.route("/api/me")
def api_me():
    user = session.get('user')
    return jsonify({"user": user, "logged_in": bool(user)})

@app.route("/api/register", methods=["POST"])
def api_register():
    if not FIREBASE_ENABLED:
        return jsonify({"error": "Firebase 미설정 — 회원가입 불가"}), 503
    data     = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '')
    if not username or not password:
        return jsonify({"error": "사용자명과 비밀번호를 입력해주세요."}), 400
    if len(username) < 2:
        return jsonify({"error": "사용자명은 2자 이상이어야 합니다."}), 400
    if len(password) < 6:
        return jsonify({"error": "비밀번호는 6자 이상이어야 합니다."}), 400
    try:
        ref = db.collection('users').document(username)
        if ref.get().exists:
            return jsonify({"error": "이미 사용 중인 사용자명입니다."}), 409
        ref.set({
            'username':      username,
            'password_hash': generate_password_hash(password),
            'created_at':    fb_fs.SERVER_TIMESTAMP,
        })
        session['user'] = username
        return jsonify({"success": True, "user": username})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/login", methods=["POST"])
def api_login():
    data     = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '')
    if not username or not password:
        return jsonify({"error": "사용자명과 비밀번호를 입력해주세요."}), 400

    # 1) Firestore 사용자 우선 확인
    if FIREBASE_ENABLED:
        try:
            doc = db.collection('users').document(username).get()
            if doc.exists:
                user_data = doc.to_dict()
                if check_password_hash(user_data.get('password_hash', ''), password):
                    session['user'] = username
                    return jsonify({"success": True, "user": username})
                return jsonify({"error": "비밀번호가 틀렸습니다."}), 401
        except Exception as e:
            return jsonify({"error": f"DB 오류: {e}"}), 500

    # 2) 환경변수 APP_USERS_JSON 폴백
    if APP_USERS.get(username) == password:
        session['user'] = username
        return jsonify({"success": True, "user": username})

    return jsonify({"error": "아이디 또는 비밀번호가 틀렸습니다."}), 401

@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.pop('user', None)
    return jsonify({"success": True})

# API 사용량 조회
@app.route("/api/usage")
def get_usage():
    if not FIREBASE_ENABLED:
        return jsonify({"count": 0, "limit": DAILY_LIMIT, "firebase": False})
    try:
        doc = db.collection('api_usage').document(_today()).get()
        count = int(doc.to_dict().get('count', 0)) if doc.exists else 0
        return jsonify({"count": count, "limit": DAILY_LIMIT, "firebase": True})
    except Exception as e:
        return jsonify({"count": 0, "limit": DAILY_LIMIT, "error": str(e)})

# 진행상황 조회
@app.route("/api/progress")
def get_progress():
    keyword = request.args.get('keyword', '').strip()
    source  = request.args.get('source', 'blog')
    if not keyword or not FIREBASE_ENABLED:
        return jsonify({"completed_pages": []})
    try:
        key = _progress_key(keyword, source)
        doc = db.collection('search_progress').document(key).get()
        completed = doc.to_dict().get('completed_pages', []) if doc.exists else []
        return jsonify({"completed_pages": completed})
    except Exception as e:
        return jsonify({"completed_pages": [], "error": str(e)})

# ── 연도 probe (display=1, 총 갯수만) ────────────────────────────────────
@app.route("/api/probe-year")
def probe_year():
    keyword = request.args.get('keyword', '').strip()
    source  = request.args.get('source', 'blog')
    year    = request.args.get('year', '')
    if not keyword or not year:
        return jsonify({"total": 0}), 400
    if not NAVER_CLIENT_ID or not NAVER_CLIENT_SECRET:
        return jsonify({"total": 0}), 500

    naver_headers = {
        "X-Naver-Client-Id":     NAVER_CLIENT_ID,
        "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
    }
    query    = f"{keyword} 직구 후기 {year}년"
    endpoint = "blog" if source == "blog" else "cafearticle"
    _, total = naver_search(endpoint, query, 1, 1, naver_headers)
    increment_usage(1)
    return jsonify({"total": total})


# ── 프로젝트 CRUD ─────────────────────────────────────────────────────────
@app.route("/api/projects", methods=["GET"])
def get_projects():
    if not FIREBASE_ENABLED:
        return jsonify({"items": [], "firebase": False})
    user = session.get('user')
    if not user:
        return jsonify({"error": "로그인 필요"}), 401
    try:
        docs = db.collection('projects').where('user_id', '==', user).stream()
        items = []
        for doc in docs:
            d = doc.to_dict()
            d['id'] = doc.id
            if 'created_at' in d and hasattr(d['created_at'], 'isoformat'):
                d['created_at'] = d['created_at'].isoformat()
            items.append(d)
        items.sort(key=lambda x: x.get('created_at') or '', reverse=True)
        return jsonify({"items": items})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/projects", methods=["POST"])
def create_project():
    if not FIREBASE_ENABLED:
        return jsonify({"error": "Firebase 미설정"}), 503
    user = session.get('user')
    if not user:
        return jsonify({"error": "로그인 필요"}), 401
    data = request.get_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({"error": "프로젝트 이름을 입력해주세요."}), 400

    keyword = name
    search_query = f"{keyword} 직구 후기"
    naver_headers = {
        "X-Naver-Client-Id":     NAVER_CLIENT_ID,
        "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
    }

    # 연도별 검색 방식이므로 총 갯수 확인만 (2 calls)
    blog_first, total_blog = naver_search("blog",        search_query, 5, 1, naver_headers)
    cafe_first, total_cafe = naver_search("cafearticle", search_query, 5, 1, naver_headers)
    increment_usage(2)

    try:
        ref = db.collection('projects').document()
        ref.set({
            'name':       name,
            'keyword':    keyword,
            'user_id':    user,
            'total_blog': total_blog,
            'total_cafe': total_cafe,
            'created_at': fb_fs.SERVER_TIMESTAMP,
        })
        doc = ref.get()
        d = doc.to_dict()
        d['id'] = doc.id
        if 'created_at' in d and hasattr(d['created_at'], 'isoformat'):
            d['created_at'] = d['created_at'].isoformat()
        return jsonify({"project": d})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/projects/<project_id>/history", methods=["PATCH"])
def update_project_history(project_id):
    if not FIREBASE_ENABLED:
        return jsonify({"ok": False})
    user = session.get('user')
    if not user:
        return jsonify({"error": "로그인 필요"}), 401
    data = request.get_json()
    # data: {source: "blog", year: 2023, months: [10, 11, 12]}
    source = data.get('source')
    year   = str(data.get('year'))
    months = data.get('months', [])
    if not source or not year or not months:
        return jsonify({"ok": False}), 400
    try:
        ref = db.collection('projects').document(project_id)
        doc = ref.get()
        if not doc.exists or doc.to_dict().get('user_id') != user:
            return jsonify({"error": "권한 없음"}), 403
        history = doc.to_dict().get('history', {})
        src_hist = history.get(source, {})
        existing = set(src_hist.get(year, []))
        existing.update(months)
        src_hist[year] = sorted(existing)
        history[source] = src_hist
        ref.update({'history': history})
        return jsonify({"ok": True, "history": history})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/projects/<project_id>", methods=["DELETE"])
def delete_project(project_id):
    if not FIREBASE_ENABLED:
        return jsonify({"error": "Firebase 미설정"}), 503
    user = session.get('user')
    if not user:
        return jsonify({"error": "로그인 필요"}), 401
    try:
        ref = db.collection('projects').document(project_id)
        doc = ref.get()
        if not doc.exists:
            return jsonify({"error": "프로젝트 없음"}), 404
        if doc.to_dict().get('user_id') != user:
            return jsonify({"error": "권한 없음"}), 403
        ref.delete()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# 선택 항목 저장 (네이버 쇼핑 조회 포함)
@app.route("/api/save-selected", methods=["POST"])
def save_selected():
    if not FIREBASE_ENABLED:
        return jsonify({
            "error": "Firebase 미설정 상태입니다. Railway Variables에 FIREBASE_SERVICE_ACCOUNT_JSON을 추가해주세요."
        }), 503
    data    = request.get_json()
    items   = data.get('items', [])
    keyword = data.get('keyword', '')
    mode    = data.get('mode', 'review')
    if not items:
        return jsonify({"error": "선택된 항목이 없습니다."}), 400

    naver_headers = {
        "X-Naver-Client-Id":     NAVER_CLIENT_ID,
        "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
    }

    # ① Vision 분석으로 영문 제품명 보강 (썸네일 있고 영문명 없는 항목만)
    for item in items:
        if item.get('product_name_en'):
            continue
        thumbnail = item.get('thumbnail') or ''
        if thumbnail and thumbnail.startswith('http'):
            vision = analyze_image_for_product(thumbnail)
            if vision.get('product_name_en'):
                item['product_name_en'] = vision['product_name_en']
            if vision.get('brand_name') and not item.get('brand_name'):
                item['brand_name'] = vision['brand_name']

    # ② 영문명 우선으로 네이버 쇼핑 조회
    shopping_cache = {}
    shopping_calls = 0
    for item in items:
        brand   = item.get('brand_name')
        product = item.get('product_name_en') or item.get('product_name')
        parts   = [p for p in [brand, product] if p]
        if not parts:
            continue
        key = " ".join(parts)
        if key not in shopping_cache:
            shopping_cache[key] = search_naver_shopping(key, naver_headers, brand_name=brand)
            shopping_calls += 1
    increment_usage(shopping_calls)

    try:
        batch = db.batch()
        for item in items:
            brand   = item.get('brand_name')
            product = item.get('product_name_en') or item.get('product_name')
            parts   = [p for p in [brand, product] if p]
            shopping_info = shopping_cache.get(" ".join(parts)) if parts else None
            ref = db.collection('sourcing_candidates').document()
            batch.set(ref, {
                **item,
                'keyword':        keyword,
                'saved_at':       fb_fs.SERVER_TIMESTAMP,
                'status':         'pending',
                'saved_by':       session.get('user', 'anonymous'),
                'mode':           mode,
                'blog_image':     item.get('thumbnail') or None,
                'product_name_en': item.get('product_name_en') or None,
                'shopping_price': shopping_info['shopping_price'] if shopping_info else None,
                'shopping_image': shopping_info['shopping_image'] if shopping_info else None,
                'shopping_link':  shopping_info['shopping_link']  if shopping_info else None,
                'shopping_mall':  shopping_info['shopping_mall']  if shopping_info else None,
                'shopping_title': shopping_info['shopping_title'] if shopping_info else None,
            })
        batch.commit()
        return jsonify({"saved": len(items), "success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# OG 이미지 프록시
@app.route("/api/og-image")
def og_image():
    url = request.args.get("url", "")
    if not url or not url.startswith("http"):
        return jsonify({"images": []})
    try:
        # 네이버 블로그: blog.naver.com → m.blog.naver.com (iframe 껍데기 우회)
        fetch_url = url
        m_blog = re.match(r'https?://blog\.naver\.com/([^/?#]+)/([0-9]+)', url)
        if m_blog:
            fetch_url = f"https://m.blog.naver.com/{m_blog.group(1)}/{m_blog.group(2)}"

        resp = requests.get(fetch_url, timeout=6, headers={
            "User-Agent": "Mozilla/5.0 (Linux; Android 10) AppleWebKit/537.36 Chrome/120.0 Mobile Safari/537.36",
            "Accept-Language": "ko-KR,ko;q=0.9",
            "Referer": "https://m.blog.naver.com/",
        })
        images = []

        # 1) og:image
        for pattern in [
            r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)["\']',
            r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image["\']',
        ]:
            for m in re.finditer(pattern, resp.text, re.IGNORECASE):
                img = m.group(1).strip()
                if img and img not in images:
                    images.append(img)

        # 2) 본문 이미지 (네이버 블로그 CDN 도메인)
        for m in re.finditer(
            r'<img[^>]+src=["\']([^"\']*(?:pstatic\.net|blogfiles\.naver\.net)[^"\']*)["\']',
            resp.text, re.IGNORECASE
        ):
            img = m.group(1).strip()
            if img.startswith('//'):
                img = 'https:' + img
            if img and img not in images:
                images.append(img)
            if len(images) >= 6:
                break

        return jsonify({"images": images})
    except Exception:
        return jsonify({"images": []})


class ExtractedProduct(BaseModel):
    brand_name:      Optional[str] = None
    product_name:    Optional[str] = None   # 한국어 상품명
    product_name_en: Optional[str] = None   # 영문 공식명/모델번호
    price_paid:      Optional[str] = None   # 후기 언급 가격
    category:        Optional[str] = None

class AllProductsResult(BaseModel):
    products: List[ExtractedProduct]

# ── AI 소싱 도우미 모델 ────────────────────────────────────────────────────
class KeywordItem(BaseModel):
    keyword: str
    category: str
    reason: str
    search_query: str
    potential: str  # "높음" / "중간" / "낮음"

class AIKeywordResult(BaseModel):
    keywords: List[KeywordItem]
    trend_summary: str

class NicheItem(BaseModel):
    item_name_ko: str
    item_name_en: str
    category: str
    sourcing_country: str   # 예: "아마존 미국", "알리익스프레스", "라쿠텐 일본", "이베이"
    sourcing_reason: str
    search_keyword: str

class NicheIdeaResult(BaseModel):
    scenario: str
    items: List[NicheItem]
    tips: str

@app.route("/api/extract-all-products", methods=["POST"])
def extract_all_products():
    """블로그 풀텍스트에서 언급된 모든 개별 상품 추출 (하울 포스트 대응)"""
    data = request.get_json()
    url  = data.get('url', '').strip()
    if not url or not url.startswith('http'):
        return jsonify({"error": "유효하지 않은 URL"}), 400
    try:
        fetch_url = url
        m_blog = re.match(r'https?://blog\.naver\.com/([^/?#]+)/([0-9]+)', url)
        if m_blog:
            fetch_url = f"https://m.blog.naver.com/{m_blog.group(1)}/{m_blog.group(2)}"

        resp = requests.get(fetch_url, timeout=8, headers={
            "User-Agent": "Mozilla/5.0 (Linux; Android 10) AppleWebKit/537.36 Chrome/120.0 Mobile Safari/537.36",
            "Accept-Language": "ko-KR,ko;q=0.9",
            "Referer": "https://m.blog.naver.com/",
        })
        # HTML → 텍스트 (스크립트/스타일 제거)
        text = re.sub(r'<script[^>]*>.*?</script>', ' ', resp.text, flags=re.DOTALL | re.IGNORECASE)
        text = re.sub(r'<style[^>]*>.*?</style>',  ' ', text,      flags=re.DOTALL | re.IGNORECASE)
        text = strip_html(text)
        text = re.sub(r'\s{3,}', '\n', text).strip()
        # 본문 앞 2000자 + 중간 2000자 조합 (최대 4000자)
        content = text[:4000] if len(text) <= 4000 else text[:2000] + '\n...\n' + text[len(text)//2:len(text)//2+2000]

        response = claude.messages.parse(
            model="claude-haiku-4-5-20251001",
            max_tokens=4000,
            messages=[{"role": "user", "content": f"""다음은 해외직구 후기 블로그 본문입니다.
본문에 등장하는 모든 개별 상품을 각각 추출해주세요.
하나의 포스트에 여러 제품이 나오면 각각 별도 항목으로 추출하세요.

추출 항목:
- brand_name: 브랜드명 (없으면 null)
- product_name: 한국어 상품명 (없으면 null)
- product_name_en: 본문에 영문으로 표기된 공식 제품명/모델번호. 영문명이 없으면 한국어 상품명을 Amazon 검색에 적합한 영어로 번역해서 입력 (필수, null 금지)
- price_paid: 후기에서 언급된 구매가격 (예: "3.98€", "$29", 없으면 null)
- category: 신발/의류/전자제품/가방/화장품/식품/기타 중 하나

블로그 본문:
{content}"""}],
            output_format=AllProductsResult,
        )
        products = [p.model_dump() for p in response.parsed_output.products]
        return jsonify({"products": products, "count": len(products)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# 메인 검색 (cursor 기반 무한스크롤)
@app.route("/api/search", methods=["POST"])
def search():
    data           = request.get_json()
    keyword        = data.get("keyword", "").strip()
    source         = data.get("source", "blog")
    mode           = data.get("mode", "review")
    year_hint      = data.get("year_hint")        # 연도 (int or None)
    start_date_str = data.get("start_date", "")
    end_date_str   = data.get("end_date", "")
    exclude_raw    = data.get("exclude_keywords", "")
    cursor         = max(1, int(data.get("cursor", 1)))

    if not keyword:
        return jsonify({"error": "키워드를 입력해주세요."}), 400
    if not NAVER_CLIENT_ID or not NAVER_CLIENT_SECRET:
        return jsonify({"error": "네이버 API 키가 설정되지 않았습니다."}), 500

    start_date = end_date = None
    try:
        if start_date_str: start_date = date.fromisoformat(start_date_str)
        if end_date_str:   end_date   = date.fromisoformat(end_date_str)
    except ValueError:
        return jsonify({"error": "날짜 형식 오류 (YYYY-MM-DD)"}), 400

    naver_headers = {
        "X-Naver-Client-Id":     NAVER_CLIENT_ID,
        "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
    }
    if mode == "gonggu":
        search_query = f"{keyword} 직구 공구"
        if source == "blog":
            source = "cafe"
    else:
        # year_hint가 있으면 연도를 쿼리에 포함 → 연도별 독립적인 1000개 풀
        if year_hint:
            search_query = f"{keyword} 직구 후기 {year_hint}년"
        else:
            search_query = f"{keyword} 직구 후기"

    MAX_NAVER_PAGE = 10
    DISPLAY        = 100
    TARGET         = 50   # 한 번 응답에 담을 최대 filtered 항목 수

    all_items  = []
    total_blog = total_cafe = 0
    next_cursor = cursor
    has_more    = False
    hit_too_old = False
    api_calls   = 0

    # MAX_PAGES_PER_CALL 없음 → 한 번 서버 호출로 10페이지 전부 스캔 (클라이언트 재호출 최소화)
    while (len(all_items) < TARGET
           and not hit_too_old
           and next_cursor <= MAX_NAVER_PAGE):

        start_pos  = (next_cursor - 1) * DISPLAY + 1
        page_items = []

        if source in ("blog", "all"):
            items, tb = naver_search("blog", search_query, DISPLAY, start_pos, naver_headers)
            for i in items: i["_source"] = "블로그"
            page_items.extend(items)
            if tb: total_blog = tb
            api_calls += 1

        if source in ("cafe", "all"):
            items, tc = naver_search("cafearticle", search_query, DISPLAY, start_pos, naver_headers)
            for i in items: i["_source"] = "카페"
            page_items.extend(items)
            if tc: total_cafe = tc
            api_calls += 1

        if not page_items:
            break

        for item in page_items:
            item["_date"] = parse_item_date(item)
        page_items.sort(key=lambda x: x["_date"] or date.min, reverse=True)

        for item in page_items:
            d = item["_date"]
            if end_date and d and d > end_date:
                continue          # 너무 최신 → skip
            if start_date and d and d < start_date:
                hit_too_old = True
                break             # 범위 이전 → 중단
            all_items.append(item)

        next_cursor += 1

    increment_usage(api_calls)

    if not hit_too_old:
        total = (total_blog if source == "blog"
                 else total_cafe if source == "cafe"
                 else total_blog + total_cafe)
        has_more = next_cursor <= MAX_NAVER_PAGE and (next_cursor - 1) * DISPLAY < total

    keyword_lower = keyword.lower()

    # 날짜필터만 적용된 원본 보존 (비교용)
    raw_items = sorted(all_items, key=lambda x: x["_date"] or date.min, reverse=True)

    # 토큰 AND 매칭: 키워드의 각 단어가 모두 제목+본문에 존재해야 함
    # "백노이즈 머신" → "백노이즈"와 "머신" 둘 다 있어야 통과 (러닝머신 등 제외)
    keyword_tokens = [t for t in keyword_lower.split() if len(t) >= 2]
    def token_match(item):
        full = strip_html(item.get("title","") + " " + item.get("description","")).lower()
        return all(t in full for t in keyword_tokens) if keyword_tokens else True
    all_items = [i for i in all_items if token_match(i)]

    exclude_keywords = [k.strip() for k in exclude_raw.split(',') if k.strip()]
    if exclude_keywords:
        def not_excluded(item):
            text = (strip_html(item.get("title", "")) + " " + strip_html(item.get("description", ""))).lower()
            return not any(kw.lower() in text for kw in exclude_keywords)
        all_items = [i for i in all_items if not_excluded(i)]

    all_items.sort(key=lambda x: x["_date"] or date.min, reverse=True)

    if not all_items:
        return jsonify({
            "results": [], "total_blog": total_blog, "total_cafe": total_cafe,
            "next_cursor": next_cursor, "has_more": has_more, "keyword": keyword,
        })

    # Claude 추출
    reviews_text = ""
    for i, item in enumerate(all_items, 1):
        title       = strip_html(item.get("title", ""))
        description = strip_html(item.get("description", ""))
        reviews_text += f"[{i}] 제목: {title}\n설명: {description}\n\n"

    if mode == "gonggu":
        claude_content = f"""다음은 카페 해외 직구 공구 관련 게시물 목록입니다. 각 게시물에서 정보를 추출해주세요.

추출 항목:
- brand_name: 브랜드명 (예: Nike, Apple, Zara 등, 없으면 null)
- product_name: 구체적인 상품명/모델명 (없으면 null)
- category: 신발/의류/전자제품/가방/화장품/식품/기타 중 하나 (없으면 null)
- purchase_source: 공구 진행 카페명 또는 구매처 (없으면 null)
- price_paid: 공구 가격 또는 구매 가격 (예: "$120", "15만원", 없으면 null)
- is_direct_purchase_review: 실제 공구 모집/진행 게시물이면 true, 단순 언급이나 후기이면 false

게시물 목록:
{reviews_text}
index는 게시물 번호 숫자를 그대로 사용하세요."""
    else:
        claude_content = f"""다음은 "{keyword}" 관련 해외 직구 후기 검색 결과입니다. 각 후기에서 정보를 추출해주세요.

추출 항목:
- brand_name: 브랜드명 (예: Nike, Apple, Zara 등, 없으면 null)
- product_name: 상품명/모델명 한국어로 (없으면 null)
- product_name_en: 본문에 영문으로 실제 표기된 공식 제품명/모델번호 (한국어를 영어로 번역하지 말 것, 텍스트에 영문이 명시된 경우에만 추출, 없으면 null)
- category: 신발/의류/전자제품/가방/화장품/식품/기타 중 하나 (없으면 null)
- purchase_source: 구매처 (아마존/이베이/알리익스프레스/직접구매/구매대행 등, 없으면 null)
- price_paid: 후기에서 언급된 구매 가격 (예: "$120", "15만원", "89달러", 없으면 null)
- is_direct_purchase_review: 아래 두 조건을 모두 만족하면 true, 하나라도 아니면 false
  조건1) 실제로 해외직구(아마존/이베이/알리/직접구매 등)로 구매한 후기일 것
  조건2) 검색 키워드 "{keyword}"와 직접 관련된 제품의 후기일 것
  (예: "{keyword}" 검색인데 전혀 다른 제품 후기면 false)

후기 목록:
{reviews_text}
index는 후기 번호 숫자를 그대로 사용하세요."""

    try:
        response = claude.messages.parse(
            model="claude-haiku-4-5-20251001",
            max_tokens=16000,
            messages=[{"role": "user", "content": claude_content}],
            output_format=ExtractionResult,
        )
    except anthropic.APIError as e:
        return jsonify({"error": f"Claude API 오류: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"error": f"분석 처리 오류: {str(e)}"}), 500

    extracted_map = {item.index: item for item in response.parsed_output.items}

    results = []
    for i, item in enumerate(all_items, 1):
        ext = extracted_map.get(i)
        results.append({
            "index":                     i,
            "source":                    item.get("_source", "블로그"),
            "title":                     strip_html(item.get("title", "")),
            "description":               strip_html(item.get("description", "")),
            "link":                      item.get("link", ""),
            "thumbnail":                 item.get("thumbnail", "") or "",
            "author":                    item.get("bloggerName") or item.get("cafename") or "",
            "postdate":                  format_date(item["_date"]),
            "brand_name":                ext.brand_name               if ext else None,
            "product_name":              ext.product_name             if ext else None,
            "product_name_en":           ext.product_name_en          if ext else None,
            "category":                  ext.category                 if ext else None,
            "purchase_source":           ext.purchase_source          if ext else None,
            "price_paid":                ext.price_paid               if ext else None,
            "is_direct_purchase_review": getattr(ext, 'is_direct_purchase_review', True) if ext else True,
        })

    raw_results = [{
        "title":    strip_html(i.get("title", "")),
        "link":     i.get("link", ""),
        "postdate": format_date(i["_date"]),
        "source":   i.get("_source", "블로그"),
    } for i in raw_items]

    return jsonify({
        "results":     results,
        "raw_results": raw_results,
        "total_blog":  total_blog,
        "total_cafe":  total_cafe,
        "next_cursor": next_cursor,
        "has_more":    has_more,
        "keyword":     keyword,
        "mode":        mode,
    })


# 소싱 후보 목록 조회
@app.route("/api/candidates")
def get_candidates():
    if not FIREBASE_ENABLED:
        return jsonify({"items": [], "firebase": False,
                        "error": "Firebase 미설정 — FIREBASE_SERVICE_ACCOUNT_JSON 환경변수를 확인해주세요."})
    try:
        user = session.get('user')
        # 로그인 유저의 항목만 조회 (saved_by 필터)
        query = db.collection('sourcing_candidates')
        if user:
            query = query.where('saved_by', '==', user)
        docs = query.limit(500).stream()
        items = []
        for doc in docs:
            d = doc.to_dict()
            d['id'] = doc.id
            # saved_at 직렬화 (Firestore timestamp → ISO string)
            if 'saved_at' in d and hasattr(d['saved_at'], 'isoformat'):
                d['saved_at'] = d['saved_at'].isoformat()
            items.append(d)
        # 최신순 정렬
        items.sort(key=lambda x: x.get('saved_at') or '', reverse=True)
        return jsonify({"items": items, "count": len(items), "firebase": True})
    except Exception as e:
        return jsonify({"items": [], "error": str(e), "firebase": True})


# 소싱 후보 상태 업데이트
ALLOWED_UPDATE_FIELDS = {
    'status', 'price_eur', 'exchange_rate', 'shipping_fee',
    'cost_price', 'margin', 'margin_rate', 'memo',
    'completed_by', 'completed_at',
    'weight_kg', 'vat_type', 'product_url', 'product_title_url',
}

@app.route("/api/candidates/<doc_id>", methods=["PATCH"])
def update_candidate(doc_id):
    if not FIREBASE_ENABLED:
        return jsonify({"error": "Firebase 미설정"}), 503
    data = request.get_json()
    update_data = {k: v for k, v in data.items() if k in ALLOWED_UPDATE_FIELDS}
    if not update_data:
        return jsonify({"error": "변경할 데이터 없음"}), 400
    try:
        db.collection('sourcing_candidates').document(doc_id).update(update_data)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# 소싱 후보 삭제
@app.route("/api/candidates/<doc_id>", methods=["DELETE"])
def delete_candidate(doc_id):
    if not FIREBASE_ENABLED:
        return jsonify({"error": "Firebase 미설정"}), 503
    try:
        db.collection('sourcing_candidates').document(doc_id).delete()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/exchange-rate")
def exchange_rate():
    """외화/KRW 최신 환율 조회 (?currency=EUR|USD|GBP|JPY|CNY|AUD|CAD)"""
    currency = request.args.get("currency", "EUR").upper()
    if currency not in {"EUR", "USD", "GBP", "JPY", "CNY", "AUD", "CAD"}:
        currency = "EUR"
    defaults = {"EUR": 1450, "USD": 1350, "GBP": 1680, "JPY": 9, "CNY": 185, "AUD": 850, "CAD": 990}
    try:
        resp = requests.get(
            f"https://api.frankfurter.app/latest?base={currency}&symbols=KRW", timeout=5
        )
        data = resp.json()
        rate = round(data["rates"]["KRW"])
        return jsonify({"rate": rate, "base": currency, "target": "KRW"})
    except Exception as e:
        return jsonify({"rate": defaults.get(currency, 1350), "base": currency, "error": str(e)})


@app.route("/api/generate-product-name", methods=["POST"])
def generate_product_name():
    """SEO 최적화 등록상품명 생성"""
    data              = request.get_json()
    brand             = data.get('brand_name', '')
    product           = data.get('product_name', '')
    product_en        = data.get('product_name_en', '')
    category          = data.get('category', '')
    country           = data.get('country', '')
    product_url       = data.get('product_url', '')
    product_title_url = data.get('product_title_url', '')  # Amazon/idealo에서 추출한 실제 제품 타이틀 (PRIMARY)
    review_title      = data.get('review_title', '')       # 블로그/카페 후기 제목 (SEO 키워드 보조)
    review_desc       = data.get('review_description', '') # 후기 본문 발췌 (SEO 키워드 보조)

    # 소싱국가에서 직구 태그 결정
    sourcing_tag = ''
    if '독일' in country or 'EU' in country:
        sourcing_tag = '독일직구'
    elif '미국' in country:
        sourcing_tag = '미국직구'
    elif '영국' in country:
        sourcing_tag = '영국직구'
    elif '프랑스' in country:
        sourcing_tag = '프랑스직구'
    elif '이탈리아' in country:
        sourcing_tag = '이탈리아직구'
    elif country and country != '-':
        sourcing_tag = '유럽직구'

    # 원문에서 숫자+단위 추출 (반드시 포함 강제)
    must_include = ''
    if product_title_url:
        nums = re.findall(r'[\d]+(?:[.,]\d+)?\s*(?:kg|g|ml|l|L|mg|oz|lb|pack|Pack|x\s*\d+|개|매|장|세트)', product_title_url, re.IGNORECASE)
        if nums:
            must_include = f"\n반드시 포함할 숫자/스펙: {' / '.join(nums)}"

    if product_title_url:
        prompt = f"""제품 타이틀을 한국어로 번역하고, 뒤에 SEO 키워드를 붙여 두 버전을 만드세요.

원문: {product_title_url}
소싱 태그: {sourcing_tag}
후기 감성 참고: {review_title}{must_include}

50자: 원문 번역 + {sourcing_tag} / 50자 이하
100자: 원문 번역 + {sourcing_tag} + 후기 감성단어 1~2개 / 100자 이하
글자수 기준: 한글·영문·숫자·공백 모두 1자 / 특수문자 금지

아래 형식으로만 출력:
50자: 결과
100자: 결과"""
    else:
        prompt = f"""네이버 스마트스토어 등록용 한국어 상품명 두 버전을 만드세요.

브랜드: {brand} / 영문명: {product_en} / 한국명: {product}
카테고리: {category} / 소싱 태그: {sourcing_tag}
후기 감성 참고: {review_title}

50자: 브랜드+제품명+소싱태그 / 50자 이하
100자: 50자 내용 + 감성키워드 1~2개 / 100자 이하
글자수 기준: 한글·영문·숫자·공백 모두 1자

아래 형식으로만 출력:
50자: 결과
100자: 결과"""

    try:
        response = claude.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=400,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = response.content[0].text.strip()
        name_50 = name_100 = ''
        for line in raw.splitlines():
            if line.startswith('50자:'):
                name_50  = line[len('50자:'):].strip()
            elif line.startswith('100자:'):
                name_100 = line[len('100자:'):].strip()
        # fallback: 파싱 실패 시 전체 텍스트를 100자로
        if not name_100:
            name_100 = raw.split('\n')[0].strip()
        if not name_50:
            name_50  = name_100[:50]
        return jsonify({"name_50": name_50, "name_100": name_100})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


ALLOWED_REG_FIELDS = {
    'product_name_display', 'name_50', 'name_100',
    'naver_price', 'coupang_price',
    'customs_rate', 'fta', 'fta_agreement', 'status', 'memo', 'country',
    'product_url', 'product_title_url', 'price_eur', 'exchange_rate',
    'weight_kg', 'shipping_fee', 'vat_type', 'customs_amt', 'vat10_amt',
    'naver_category', 'reg_images',
}

@app.route("/api/product-registrations", methods=["GET"])
def get_product_registrations():
    if not FIREBASE_ENABLED:
        return jsonify({"items": [], "firebase": False})
    try:
        docs = db.collection('product_registrations').limit(500).stream()
        items = []
        for doc in docs:
            d = doc.to_dict(); d['id'] = doc.id
            if 'created_at' in d and hasattr(d['created_at'], 'isoformat'):
                d['created_at'] = d['created_at'].isoformat()
            items.append(d)
        items.sort(key=lambda x: x.get('created_at') or '', reverse=True)
        return jsonify({"items": items})
    except Exception as e:
        return jsonify({"items": [], "error": str(e)})


@app.route("/api/product-registrations", methods=["POST"])
def create_product_registration():
    if not FIREBASE_ENABLED:
        return jsonify({"error": "Firebase 미설정"}), 503
    items = request.get_json().get('items', [])
    if not items:
        return jsonify({"error": "항목 없음"}), 400
    try:
        batch = db.batch()
        for item in items:
            ref = db.collection('product_registrations').document()
            batch.set(ref, {**item, 'created_at': fb_fs.SERVER_TIMESTAMP,
                            'created_by': session.get('user', 'anonymous'), 'status': 'draft'})
        batch.commit()
        return jsonify({"created": len(items)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/product-registrations/<doc_id>", methods=["PATCH"])
def update_product_registration(doc_id):
    if not FIREBASE_ENABLED:
        return jsonify({"error": "Firebase 미설정"}), 503
    data = {k: v for k, v in request.get_json().items() if k in ALLOWED_REG_FIELDS}
    if not data:
        return jsonify({"error": "변경할 데이터 없음"}), 400
    try:
        db.collection('product_registrations').document(doc_id).update(data)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/product-registrations/<doc_id>", methods=["DELETE"])
def delete_product_registration(doc_id):
    if not FIREBASE_ENABLED:
        return jsonify({"error": "Firebase 미설정"}), 503
    try:
        db.collection('product_registrations').document(doc_id).delete()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


CLOUDINARY_CLOUD = os.environ.get('CLOUDINARY_CLOUD_NAME', '')
CLOUDINARY_PRESET = os.environ.get('CLOUDINARY_UPLOAD_PRESET', '')

def _upload_image_to_storage(img_url: str, doc_id: str, idx: int = 0) -> str:
    """외부 이미지 URL을 Firebase Storage에 업로드하고 공개 URL 반환.
    Firebase 미설정 시 Cloudinary 시도, 둘 다 없으면 원본 URL 반환."""
    if not img_url or not img_url.startswith('http'):
        return img_url

    # ── 1) Firebase Storage 업로드 (우선)
    if FIREBASE_ENABLED and STORAGE_BUCKET:
        try:
            img_resp = requests.get(img_url, headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Accept': 'image/webp,image/apng,image/*,*/*'
            }, timeout=15)
            if img_resp.status_code == 200:
                content_type = img_resp.headers.get('Content-Type', 'image/jpeg').split(';')[0].strip()
                ext = 'jpg' if 'jpeg' in content_type or 'jpg' in content_type else content_type.split('/')[-1]
                blob_path = f"sourcing/{doc_id}/img_{idx}.{ext}"
                bucket = fb_storage.bucket()
                blob = bucket.blob(blob_path)
                blob.upload_from_string(img_resp.content, content_type=content_type)
                blob.make_public()
                pub_url = blob.public_url
                print(f"[Firebase Storage] 업로드 성공: {pub_url[:80]}")
                return pub_url
        except Exception as e:
            print(f"[Firebase Storage] 오류 {img_url[:60]}: {e}")

    # ── 2) Cloudinary (Firebase 실패 시)
    if CLOUDINARY_CLOUD and CLOUDINARY_PRESET:
        try:
            folder = f"sourcing/{doc_id}"
            resp = requests.post(
                f"https://api.cloudinary.com/v1_1/{CLOUDINARY_CLOUD}/image/upload",
                data={
                    'file': img_url,
                    'upload_preset': CLOUDINARY_PRESET,
                    'folder': folder,
                    'public_id': f"img_{idx}",
                    'overwrite': 'true',
                },
                timeout=30,
            )
            result = resp.json()
            if 'secure_url' in result:
                url = result['secure_url']
                url = url.replace('/upload/', '/upload/c_pad,b_white,w_1000,h_1000,f_jpg/')
                return url
            print(f"[Cloudinary] 업로드 실패: {result.get('error', result)}")
        except Exception as e:
            print(f"[Cloudinary] 오류 {img_url[:60]}: {e}")

    return img_url


def _fetch_product_page_data(url: str):
    """제품 URL에서 이미지 목록 + 본문 텍스트 추출"""
    try:
        fetch_url = url
        if 'amazon.' in url and 'language=' not in url:
            sep = '&' if '?' in url else '?'
            fetch_url = url + sep + 'language=en_US'  # 영어로 받아서 Claude가 번역
        resp = requests.get(fetch_url, timeout=10, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0 Safari/537.36",
            "Accept-Language": "en-US,en;q=0.9",
            "Accept": "text/html,application/xhtml+xml",
        })
        html = resp.text

        images = []

        # ── Amazon 이미지 ──
        if 'amazon.' in url:
            # 메인 이미지: data-a-dynamic-image JSON
            m = re.search(r'data-a-dynamic-image=["\']([^"\']+)["\']', html)
            if m:
                raw = m.group(1).replace('&quot;', '"')
                try:
                    img_dict = json.loads(raw)
                    # 가장 큰 이미지 우선 (width 기준)
                    sorted_imgs = sorted(img_dict.items(), key=lambda x: x[1][0] if x[1] else 0, reverse=True)
                    images += [k for k, _ in sorted_imgs]
                except Exception:
                    pass
            # 보조 이미지: altImages 썸네일 URL → 고해상도로 변환
            for m2 in re.finditer(r'https://m\.media-amazon\.com/images/I/[A-Za-z0-9%+_-]+\._[A-Z0-9_,]+_\.(jpg|png|jpeg)', html):
                large = re.sub(r'\._[A-Z0-9_,]+_\.', '.', m2.group(0))
                if large not in images:
                    images.append(large)
        else:
            # ── 일반 페이지: <img> src 수집 ──
            for m3 in re.finditer(r'<img[^>]+(?:src|data-src)=["\']([^"\']+)["\']', html, re.IGNORECASE):
                src = m3.group(1)
                if src.startswith('http') and any(ext in src.lower() for ext in ['.jpg', '.jpeg', '.png', '.webp']):
                    if not any(x in src.lower() for x in ['logo', 'icon', 'sprite', 'banner', 'badge', 'pixel']):
                        if src not in images:
                            images.append(src)

        # 이미지 중복 제거, 최대 10개
        seen = set(); uniq = []
        for img in images:
            if img not in seen:
                seen.add(img); uniq.append(img)
        images = uniq[:10]

        # ── 본문 텍스트 추출 ──
        page_text = strip_html(html)
        # 너무 긴 경우 제품명 + 핵심 설명 부분만
        page_text = re.sub(r'\n{3,}', '\n\n', page_text).strip()[:4000]

        return images, page_text
    except Exception:
        return [], ''


def _generate_korean_description(product_name: str, page_text: str, images: list, item: dict = None) -> str:
    """Claude로 제품 페이지 텍스트를 한국어 HTML 상세설명으로 변환
    page_text가 없으면 저장된 item 필드(product_title_url 등)로 대체 생성"""
    item = item or {}

    # 이미지 HTML (상단)
    img_html = '\n'.join(
        f'<img src="{img}" style="max-width:860px;display:block;margin:8px auto">'
        for img in images[:9]
    )

    # 텍스트 소스 결정: 페이지 직접 fetch > 저장된 product_title_url > 기본 필드
    if not page_text:
        product_title_url = item.get('product_title_url', '')
        brand = item.get('brand_name', '')
        category = item.get('category', '')
        country = item.get('country', '')
        product_name_en = item.get('product_name_en', '') or item.get('product_name', '')
        sourcing_url = item.get('product_url', '')

        if product_title_url:
            page_text = f"Product Title: {product_title_url}\nBrand: {brand}\nCategory: {category}"
        elif product_name_en:
            page_text = f"Product: {product_name_en}\nBrand: {brand}\nCategory: {category}\nCountry: {country}"
        else:
            # 텍스트도 없고 저장된 정보도 없으면 이미지만
            return img_html

    try:
        prompt = f"""다음 제품 정보를 바탕으로 한국 네이버 스마트스토어 상품 상세설명용 HTML을 작성해주세요.

제품명(한국어): {product_name}
제품 정보:
{page_text[:3000]}

요구사항:
1. 제품 특징·스펙·사용방법을 한국어로 작성 (번역/요약)
2. HTML 태그 사용: <h3>, <ul>, <li>, <p>, <strong>, <table>
3. 배송·광고·법적문구 제외
4. 500~1000자 분량
5. <html>/<body> 태그 없이 내용만 반환"""

        resp = claude.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1500,
            messages=[{"role": "user", "content": prompt}],
        )
        desc_html = resp.content[0].text.strip()
        # markdown code fence 제거 (```html ... ``` 또는 ``` ... ```)
        desc_html = re.sub(r'^```(?:html)?\s*\n?', '', desc_html)
        desc_html = re.sub(r'\n?```\s*$', '', desc_html).strip()
    except Exception:
        desc_html = f'<p><strong>{product_name}</strong></p>'

    return (img_html + '\n' + desc_html) if img_html else desc_html


@app.route("/api/export-excel", methods=["POST"])
def export_excel():
    """선택된 상품등록대장 항목을 네이버 스마트스토어 일괄등록 Excel로 내보내기"""
    if not FIREBASE_ENABLED:
        return jsonify({"error": "Firebase 미설정"}), 503

    data = request.get_json()
    ids = data.get("ids", [])  # 선택된 doc ID 목록. 빈 리스트면 전체

    try:
        col_ref = db.collection('product_registrations')
        if ids:
            # Firestore는 in 쿼리 10개 제한 → 직접 fetch
            docs = [col_ref.document(doc_id).get() for doc_id in ids]
            items = [dict(d.to_dict(), id=d.id) for d in docs if d.exists]
        else:
            items = [dict(d.to_dict(), id=d.id) for d in col_ref.limit(200).stream()]
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    if not items:
        return jsonify({"error": "내보낼 항목이 없습니다"}), 400

    # 원산지코드 매핑 (ISO 3166-1 numeric 기반 네이버 코드)
    ORIGIN_CODE = {
        '독일': '0276', '미국': '0840', '일본': '0392', '중국': '0156',
        '영국': '0826', '프랑스': '0250', '이탈리아': '0380', '스페인': '0724',
        '캐나다': '0124', '호주': '0036', '네덜란드': '0528', '스위스': '0756',
        '오스트리아': '0040', '벨기에': '0056', '폴란드': '0616',
        'Germany': '0276', 'USA': '0840', 'Japan': '0392', 'China': '0156',
    }

    # 카테고리 자동 매핑 (앱 카테고리 → 네이버 카테고리코드)
    CATEGORY_CODE = {
        '식품':    50001921,  # 식품 > 과자/베이커리 > 기타과자
        '의류':    50000803,  # 패션의류 > 여성의류 > 티셔츠
        '신발':    50003839,  # 패션잡화 > 여성신발 > 운동화 > 워킹화
        '가방':    50000639,  # 패션잡화 > 여성가방 > 숄더백
        '전자제품': 50001579,  # 디지털/가전 > PC액세서리 > 기타PC액세서리
        '화장품':  50000440,  # 화장품/미용 > 스킨케어 > 크림
        '기타':    50001921,  # 식품 > 과자/베이커리 > 기타과자 (fallback)
    }

    # 템플릿 로드
    template_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Documents', 'old', 'ExcelSaveTemplate_20260309.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # 컬럼 헤더 → 인덱스 맵 (row 2 기준)
    header_map = {}
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=2, column=col).value
        if h:
            header_map[str(h).replace('\n', '').strip()] = col

    def col_idx(name):
        return header_map.get(name)

    # 행 3~6 (가이드) 삭제 — 단, 원산지코드 예시행(row4)의 '@' 포맷 먼저 저장
    origin_col = col_idx('원산지코드')
    origin_nf  = ws.cell(4, origin_col).number_format if origin_col else '@'
    ws.delete_rows(3, 4)

    # 데이터 행 기입 (행 3~)
    for row_num, item in enumerate(items, start=3):
        def w(col_name, val):
            idx = col_idx(col_name)
            if idx:
                ws.cell(row=row_num, column=idx).value = val

        product_name = item.get('name_100') or item.get('name_50') or item.get('product_name_display') or ''
        brand        = item.get('brand_name', '') or ''
        product_url  = item.get('product_url', '') or ''

        # ── 필수: 상품명
        w('상품명', product_name)

        # ── 필수: 판매가 (저장 값 없으면 totalCost 기반 제안가 계산)
        naver_price = item.get('naver_price')
        price_val = None
        if naver_price:
            try:
                price_val = (int(str(naver_price).replace(',', '').replace('원', '').strip()) // 10) * 10
            except Exception:
                pass
        if not price_val and item.get('price_eur'):
            import math as _math
            _rate    = float(item.get('exchange_rate') or 1450)
            _vat     = float(item.get('vat_type') or 19) / 100
            _eur     = float(item.get('price_eur') or 0)
            _eur_nv  = _eur / (1 + _vat)
            _vat_krw = round(_eur_nv * _rate)
            _ship    = float(item.get('shipping_fee') or 0)
            _cust    = float(item.get('customs_amt')  or 0)
            _vat10   = float(item.get('vat10_amt')    or 0)
            _sfee    = 5500 if _eur_nv < 1000 else round(_eur * _rate * 0.01)
            _afee    = 6000 if _eur_nv <= 30 else (8000 if _eur_nv <= 50 else 10000)
            _total   = _vat_krw + _ship + _cust + _vat10 + _sfee + _afee
            price_val = int(_math.ceil((_total + 10000) / (1 - 0.0585) / 100) * 100)
        if price_val:
            w('판매가', price_val)

        # ── 필수: 재고수량
        w('재고수량', 100)

        # ── 필수: 원산지코드 (한글 우선 추출, 이모지 안전하게 제거)
        country = item.get('country', '') or ''
        korean_m = re.search(r'[가-힣]+', country)
        if korean_m:
            country_key = korean_m.group(0)
        else:
            country_key = re.sub(r'[^a-zA-Z\s]', '', country).strip()
        origin = ORIGIN_CODE.get(country_key, '0276')  # 기본: 독일
        print(f"[Excel] 원산지: country={repr(country)} → key={repr(country_key)} → code={repr(origin)} col={origin_col}")
        if origin_col:
            c = ws.cell(row=row_num, column=origin_col)
            c.number_format = '@'
            c.value = str(origin)
            print(f"[Excel] 원산지코드 기입: row={row_num} col={origin_col} val={repr(c.value)}")

        # ── 필수: 카테고리코드 (저장된 naver_category 우선, 없으면 앱 카테고리 자동 매핑)
        naver_cat = item.get('naver_category') or item.get('naver_category_code')
        if not naver_cat:
            app_cat = item.get('category', '') or ''
            naver_cat = CATEGORY_CODE.get(app_cat)
        if naver_cat:
            w('카테고리코드', int(naver_cat))

        # ── 이미지: reg_images 우선, 없으면 쇼핑 이미지 fallback
        use_cloudinary = bool(CLOUDINARY_CLOUD and CLOUDINARY_PRESET)
        reg_images = item.get('reg_images') or []
        if not reg_images:
            reg_images = [x for x in [
                item.get('shopping_image'), item.get('blog_image'), item.get('thumbnail')
            ] if x and x.startswith('http')]
        reg_images = list(dict.fromkeys(reg_images))[:5]

        uploaded_images = []
        for idx, img_url in enumerate(reg_images):
            pub_url = _upload_image_to_storage(img_url, item.get('id', f'item_{row_num}'), idx)
            if pub_url:
                uploaded_images.append(pub_url)

        if uploaded_images:
            w('대표이미지', uploaded_images[0])
        if len(uploaded_images) > 1:
            w('추가이미지', '\n'.join(uploaded_images[1:5]))

        # ── 상세설명: Claude 한국어 번역/요약
        desc_html = _generate_korean_description(product_name, '', reg_images, item)
        w('상세설명', desc_html)

        # ── 선택 필드
        w('브랜드', brand)
        w('제조사', brand)

        # ── 부가세
        vat_type = str(item.get('vat_type', '') or '')
        if '면세' in vat_type:
            w('부가세', '면세상품')
        else:
            w('부가세', '과세상품')

        # 배송방법 + 필수 배송 필드
        w('배송방법', '택배, 소포, 등기')
        w('택배사코드', 'HANJIN')
        w('배송비유형', '무료')
        w('기본배송비', 0)
        w('배송비 결제방식', '선결제')
        w('반품배송비', 50000)
        w('교환배송비', 100000)

        # A/S
        w('A/S 전화번호', '070-4571-6921')
        w('A/S 안내', '해외 구매 대행 상품으로 A/S는 불가합니다.\n자세한 내용은 상세페이지를 참조해주세요.')

        # 관부가세
        w('관부가세', '부과 대상 아님')

        # 상품상태
        w('상품상태', '신상품')

    # 파일 저장 후 반환
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.now().strftime('%Y%m%d')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'스마트스토어_일괄등록_{today}.xlsx'
    )


@app.route("/api/fetch-weight", methods=["POST"])
def fetch_weight():
    """Amazon.de / idealo 제품 페이지에서 무게+가격+타이틀 추출"""
    req_data = request.get_json()
    url = req_data.get('url', '').strip()
    candidate_id = req_data.get('candidate_id', 'tmp')
    if not url or not url.startswith('http'):
        return jsonify({"weight": None, "error": "유효한 URL이 아닙니다"}), 400
    try:
        # Amazon URL에 language=de_DE 강제 (EUR 가격 + 독어 필드명 보장)
        fetch_url = url
        if 'amazon.' in url and 'language=' not in url:
            sep = '&' if '?' in url else '?'
            fetch_url = url + sep + 'language=de_DE'
        resp = requests.get(fetch_url, timeout=10, headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
            "Accept-Language": "de-DE,de;q=0.9,en-US;q=0.8,en;q=0.7",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Encoding": "gzip, deflate, br",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
        })
        if resp.status_code != 200:
            slug_title = None
            if 'amazon.' in url:
                slug_m = re.search(r'amazon\.[^/]+/([^/]+)/dp/', url)
                if slug_m and slug_m.group(1) != 'dp':
                    slug_title = slug_m.group(1).replace('-', ' ')
            return jsonify({"weight": None, "product_title": slug_title,
                            "error": f"페이지 로드 실패 ({resp.status_code})"})
        # Amazon 봇 차단 감지 (CAPTCHA 또는 Robot Check 페이지)
        if 'amazon.' in url and (
            'Robot Check' in resp.text or
            'Enter the characters you see below' in resp.text or
            'api-services-support@amazon.com' in resp.text or
            resp.text.strip().count('<') < 10
        ):
            # URL 슬러그에서 상품명 추출 (amazon.de/SLUG/dp/ASIN/...)
            slug_title = None
            slug_m = re.search(r'amazon\.[^/]+/([^/]+)/dp/', url)
            if slug_m:
                slug = slug_m.group(1)
                if slug and slug != 'dp':
                    slug_title = slug.replace('-', ' ')
            return jsonify({"weight": None, "product_title": slug_title,
                            "error": "Amazon 봇 차단 — 무게/가격은 직접 입력해주세요"})

        text = strip_html(resp.text)

        # 정규식으로 무게 패턴 찾기 (독어 Kilogramm + 영어 Kilograms + kg/g 모두 처리)
        weight_patterns = [
            (r'(?:Artikelgewicht|Item\s+Weight|Gewicht|Versandgewicht|Stückgewicht|Produktgewicht)[^\d]{0,30}([0-9]+[.,][0-9]*)\s*(kg|Kilogramm|Kilograms?)', 'kg'),
            (r'(?:Artikelgewicht|Item\s+Weight|Gewicht|Versandgewicht|Stückgewicht|Produktgewicht)[^\d]{0,30}([0-9]+)\s*(kg|Kilogramm|Kilograms?)', 'kg'),
            (r'(?:Artikelgewicht|Item\s+Weight|Gewicht|Versandgewicht|Stückgewicht|Produktgewicht)[^\d]{0,30}([0-9]+[.,][0-9]*)\s*(g|Gramm|Grams?)\b', 'g'),
            (r'(?:Artikelgewicht|Item\s+Weight|Gewicht|Versandgewicht|Stückgewicht|Produktgewicht)[^\d]{0,30}([0-9]+)\s*(g|Gramm|Grams?)\b', 'g'),
            # prodDetAttrValue 형식 (Amazon 스펙 테이블)
            (r'prodDetAttrValue[^>]*>\s*(?:&lrm;)?\s*([0-9]+[.,][0-9]*)\s*(Kilograms?|Kilogramm|kg)', 'kg'),
            (r'prodDetAttrValue[^>]*>\s*(?:&lrm;)?\s*([0-9]+[.,][0-9]*)\s*(Grams?|Gramm|g)\b', 'g'),
        ]
        found_weight = None
        for pattern, default_unit in weight_patterns:
            m = re.search(pattern, resp.text, re.IGNORECASE)  # HTML 원문에서 직접 검색
            if m:
                val = float(m.group(1).replace(',', '.'))
                if default_unit == 'g':
                    val = val / 1000
                found_weight = round(val, 3)
                break
        # strip_html 텍스트에서도 재시도
        if found_weight is None:
            for pattern, default_unit in weight_patterns[:4]:
                m = re.search(pattern, text, re.IGNORECASE)
                if m:
                    val = float(m.group(1).replace(',', '.'))
                    if default_unit == 'g': val = val / 1000
                    found_weight = round(val, 3)
                    break

        # 가격 정규식 (EUR)
        found_price = None
        # 1순위: corePriceDisplay div 안의 a-offscreen span (€39,99 형식)
        core_m = re.search(
            r'id=["\']corePriceDisplay_desktop_feature_div["\'][\s\S]{0,2000}?'
            r'class=["\']a-offscreen["\'][^>]*>([\s\S]{0,30}?)</span>',
            resp.text, re.IGNORECASE
        )
        if core_m:
            raw_p = re.sub(r'[^\d.,]', '', core_m.group(1))
            if raw_p:
                try:
                    candidate = round(float(raw_p.replace(',', '.')), 2)
                    if 0 < candidate < 100000:
                        found_price = candidate
                except ValueError:
                    pass
        # 2순위: a-price-whole + a-price-fraction
        if not found_price:
            ap_m = re.search(
                r'a-price-whole["\'][^>]*>([0-9]+)[^<]*</span>[\s\S]{0,100}?'
                r'a-price-fraction["\'][^>]*>([0-9]{2})',
                resp.text, re.DOTALL
            )
            if ap_m:
                candidate = float(f"{ap_m.group(1)}.{ap_m.group(2)}")
                if 0 < candidate < 100000:
                    found_price = candidate
        # 3순위: 일반 EUR 패턴
        if not found_price:
            for pp in [r'([0-9]+[.,][0-9]{2})\s*€', r'€\s*([0-9]+[.,][0-9]{2})',
                       r'EUR\s*([0-9]+[.,][0-9]{2})', r'"price":\s*"EUR ([0-9]+[.,][0-9]{2})"']:
                pm = re.search(pp, resp.text)
                if pm:
                    candidate = round(float(pm.group(1).replace(',', '.')), 2)
                    if 0 < candidate < 100000:
                        found_price = candidate
                        break

        # 소싱처 이미지 추출 (Amazon 전용)
        page_images = []
        if 'amazon.' in url:
            # 메인 이미지
            m_img = re.search(r'data-a-dynamic-image=["\']([^"\']+)["\']', resp.text)
            if m_img:
                try:
                    img_dict = json.loads(m_img.group(1).replace('&quot;', '"'))
                    base_best = {}
                    for full_url, dims in img_dict.items():
                        base = re.sub(r'\._[A-Z0-9_,]+_\.', '.', full_url)
                        w = dims[0] if dims else 0
                        if base not in base_best or w > base_best[base][1]:
                            base_best[base] = (full_url, w)
                    for u, _ in sorted(base_best.values(), key=lambda x: x[1], reverse=True):
                        page_images.append(u)
                except Exception:
                    pass
            # 서브 이미지: 썸네일 스트립에서 고해상도 변환
            seen_ids = set(re.search(r'/images/I/([^.]+)', u).group(1) for u in page_images if re.search(r'/images/I/([^.]+)', u))
            for m2 in re.finditer(r'https://m\.media-amazon\.com/images/I/([A-Za-z0-9%+_-]+)\._[A-Z0-9_,]+_\.(jpg|png|jpeg)', resp.text):
                img_id = m2.group(1)
                if img_id not in seen_ids:
                    seen_ids.add(img_id)
                    page_images.append(f'https://m.media-amazon.com/images/I/{img_id}.{m2.group(2)}')
        page_images = page_images[:5]
        # Cloudinary에 업로드해서 안정적인 CDN URL로 변환
        doc_id = candidate_id
        uploaded = []
        for i, img in enumerate(page_images):
            cdn = _upload_image_to_storage(img, doc_id, i)
            if cdn:
                uploaded.append(cdn)
        if uploaded:
            page_images = uploaded

        # 제품 타이틀 추출
        product_title = None
        # 1순위: id="productTitle" span — 열린 태그에서 500자 잘라 내부 HTML 전체 스트립
        pt_m = re.search(r'id=["\']productTitle["\'][^>]*>([\s\S]{0,500}?)</span>', resp.text, re.IGNORECASE)
        if pt_m:
            raw = re.sub(r'<[^>]+>', '', pt_m.group(1))  # HTML 태그 제거
            raw = raw.replace('&#x27;', "'").replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>').replace('&lrm;', '').replace('&#x200F;', '')
            product_title = re.sub(r'\s+', ' ', raw).strip()
            if len(product_title) < 3:
                product_title = None
        # 2순위: <title> 태그 fallback
        if not product_title:
            title_m = re.search(r'<title[^>]*>(.*?)</title>', resp.text, re.IGNORECASE | re.DOTALL)
            if title_m:
                raw_title = strip_html(title_m.group(1)).strip()
                for suffix in [': Amazon.de', ': Amazon.com', ' | Amazon', ' - idealo', ' | idealo', ' - Amazon']:
                    if suffix.lower() in raw_title.lower():
                        raw_title = raw_title[:raw_title.lower().index(suffix.lower())].strip()
                if len(raw_title) > 5:
                    product_title = raw_title

        if found_weight is not None and found_price is not None:
            return jsonify({"weight": found_weight, "price_eur": found_price, "unit": "kg",
                            "source": "regex", "product_title": product_title, "images": page_images})
        if found_weight is not None:
            pass  # 가격은 Claude로 보완 시도

        # Claude로 무게+가격 추출 (정규식에서 못 찾은 필드 보완)
        relevant_text = text[:4000]
        for kw in ['Gewicht', 'Weight', 'Artikelgewicht', 'Preis', 'Price', '€', 'kg']:
            idx = text.find(kw)
            if idx > 200:
                relevant_text = text[max(0, idx - 300): idx + 800]
                break

        resp_claude = claude.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=120,
            messages=[{"role": "user", "content":
                f'Extract product weight and price from this text. '
                f'Reply ONLY with JSON: {{"weight": 0.5, "unit": "kg", "price_eur": 29.99}}. '
                f'Weight in kg (convert grams to kg). Price in EUR as number only. '
                f'Use null for any field not found.\n\n{relevant_text}'
            }],
        )
        result_text = re.sub(r'```json?\s*|\s*```', '', resp_claude.content[0].text.strip()).strip()
        result = json.loads(result_text)
        out = {}
        if found_weight is not None:
            out['weight'] = found_weight
        elif result.get('weight') is not None:
            w = float(result['weight'])
            if result.get('unit', 'kg') in ('g', 'gram', 'Gramm'):
                w = w / 1000
            out['weight'] = round(w, 3)
        if found_price is not None:
            out['price_eur'] = found_price
        elif result.get('price_eur') is not None:
            out['price_eur'] = round(float(result['price_eur']), 2)
        if out:
            out['source'] = 'claude'
            out['product_title'] = product_title
            out['images'] = page_images
            return jsonify(out)

        if product_title:
            return jsonify({"weight": None, "product_title": product_title, "images": page_images,
                            "error": "무게/가격 정보를 찾을 수 없습니다. 직접 입력해주세요."})
        return jsonify({"weight": None, "images": page_images, "error": "무게/가격 정보를 찾을 수 없습니다. 직접 입력해주세요."})
    except Exception as e:
        return jsonify({"weight": None, "error": str(e)}), 500


# ── AI 소싱 도우미 엔드포인트 ──────────────────────────────────────────────
@app.route('/api/ai-keywords', methods=['POST'])
def api_ai_keywords():
    if not session.get('user'):
        return jsonify({"error": "로그인이 필요합니다"}), 401
    data     = request.get_json() or {}
    category = data.get('category', '전체').strip()
    custom   = data.get('custom', '').strip()

    cat_hint = f"카테고리: {category}" if category != '전체' else "카테고리: 전 분야"
    custom_hint = f"\n추가 힌트: {custom}" if custom else ""

    prompt = f"""당신은 한국 해외직구 구매대행 전문가입니다.
{cat_hint}{custom_hint}

아래 조건으로 소싱 키워드 15개를 추천하세요:
- 네이버 블로그/카페에서 직구 후기 검색 시 실제 결과가 많이 나올 키워드
- 마진율이 높고 국내가보다 해외가가 유리한 제품
- 구체적인 브랜드명/모델명 포함 권장
- keyword: 검색 핵심어 (짧게, 예: "다이슨 에어랩")
- search_query: 실제 검색에 쓸 전체 쿼리 (예: "다이슨 에어랩 직구 후기")
- potential: 소싱 매력도 ("높음"/"중간"/"낮음")
- reason: 추천 이유 (1~2문장)
- trend_summary: 전반적인 트렌드 요약 (2~3문장)

JSON 형식으로만 응답하세요."""

    try:
        result = claude.messages.parse(
            model="claude-haiku-4-5-20251001",
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}],
            output_format=AIKeywordResult,
        )
        return jsonify(result.parsed_output.model_dump())
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/ai-niche-ideas', methods=['POST'])
def api_ai_niche_ideas():
    if not session.get('user'):
        return jsonify({"error": "로그인이 필요합니다"}), 401
    data     = request.get_json() or {}
    scenario = data.get('scenario', '').strip()
    if not scenario:
        return jsonify({"error": "scenario 필드가 필요합니다"}), 400

    prompt = f"""당신은 한국 해외직구 구매대행 전문가입니다.

시나리오: "{scenario}"

이 상황에서 필요한 물품 중 해외직구로 소싱하면 유리한 아이템 20개를 추천하세요.
- 일반인이 자주 생각하지 못하는 틈새 아이템 위주
- 국내 대비 해외가가 훨씬 저렴한 것 우선
- item_name_ko: 한국어 상품명
- item_name_en: 영문명 또는 모델명
- category: 카테고리
- sourcing_country: 가장 저렴하게 살 수 있는 나라/플랫폼 (예: "아마존 미국", "알리익스프레스", "라쿠텐 일본", "이베이", "독일 아마존", "1688")
- sourcing_reason: 소싱하면 좋은 이유 (1문장)
- search_keyword: 이 제품을 해외직구로 구매한 한국인이 네이버 블로그에 후기를 쓸 때 실제로 사용하는 단어. 제품명이 아니라 '검색어' 관점에서 생각할 것. 예) 방음귀마개→"귀마개", 백노이즈머신→"수면소음기", 에스프레소머신→"커피머신". 한국어만, 짧게, 사람들이 실제로 검색하는 단어로.
- tips: 이 시나리오 소싱 전략 팁 (2~3문장)

JSON 형식으로만 응답하세요."""

    try:
        result = claude.messages.parse(
            model="claude-haiku-4-5-20251001",
            max_tokens=3000,
            messages=[{"role": "user", "content": prompt}],
            output_format=NicheIdeaResult,
        )
        data = result.parsed_output.model_dump()

        # 각 키워드를 네이버에서 실제 검색해서 블로그 건수 확인
        naver_headers = {
            "X-Naver-Client-Id":     NAVER_CLIENT_ID,
            "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
        }
        for item in data.get("items", []):
            kw = item.get("search_keyword", "")
            if kw:
                _, total = naver_search("blog", f"{kw} 직구 후기", 1, 1, naver_headers)
                item["naver_count"] = total
            else:
                item["naver_count"] = 0
        increment_usage(len(data.get("items", [])))

        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/naver-shopping-specs", methods=["POST"])
def naver_shopping_specs():
    data            = request.get_json()
    candidate_id    = data.get("candidate_id", "")
    product_name    = (data.get("product_name")    or "").strip()
    product_name_en = (data.get("product_name_en") or "").strip()
    brand_name      = (data.get("brand_name")      or "").strip()

    # 검색 쿼리: 영문명 우선, 없으면 한국어
    query = product_name_en or product_name
    if not query:
        return jsonify({"error": "상품명이 없습니다."}), 400
    if brand_name and brand_name.lower() not in query.lower():
        query = f"{brand_name} {query}"

    naver_headers = {
        "X-Naver-Client-Id":     NAVER_CLIENT_ID,
        "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
    }

    try:
        resp = requests.get(
            "https://openapi.naver.com/v1/search/shop.json",
            headers=naver_headers,
            params={"query": query, "display": 30, "sort": "asc"},
            timeout=10,
        )
        rj        = resp.json() if resp.status_code == 200 else {}
        raw_items = rj.get("items", [])
        shop_total = rj.get("total", 0)
    except Exception:
        raw_items  = []
        shop_total = 0

    increment_usage(1)

    if not raw_items:
        return jsonify({"specs": [], "product_match": query, "total": shop_total})

    items_text = ""
    for i, it in enumerate(raw_items, 1):
        title  = strip_html(it.get("title", ""))
        price  = it.get("lprice", 0)
        mall   = strip_html(it.get("mallName", ""))
        link   = it.get("link", "")
        items_text += f"[{i}] {title} | {price}원 | {mall} | {link}\n"

    try:
        response = claude.messages.parse(
            model="claude-haiku-4-5-20251001",
            max_tokens=2000,
            messages=[{"role": "user", "content": f"""네이버 쇼핑에서 '{query}' 검색 결과 {len(raw_items)}개입니다.

스펙(중량/수량/팩수)별로 그룹화하여 각 스펙의 최저가 상품 1개만 추출하세요.

규칙:
- spec: "200g 1개", "200g 2팩", "500g", "1kg 3팩" 등 명확하게
- price: 해당 스펙 중 가장 낮은 가격(원 정수)
- free_shipping: 상품명/판매처에 "무료배송", "로켓배송", "로켓직구", "무료", "free shipping" 등이 있으면 true, "배송비" 뒤에 금액이 있으면 false, 불명확하면 null
- delivery_fee: "배송비 9,900원", "배송 3,000원" 등 명시된 금액이 있으면 정수(원), 없으면 null
- seller: mallName
- link: 해당 상품 링크 그대로
- 스펙 불분명하거나 명백히 관련 없는 상품 제외
- 중복 스펙은 최저가 1개만

상품 목록:
{items_text}"""}],
            output_format=SpecResult,
        )
        specs = [
            {
                "spec":          s.spec,
                "price":         s.price,
                "free_shipping": s.free_shipping,
                "delivery_fee":  s.delivery_fee,
                "seller":        s.seller or "",
                "link":          s.link or "",
            }
            for s in response.parsed_output.specs
        ]
    except Exception as e:
        return jsonify({"error": f"Claude 분석 오류: {str(e)}"}), 500

    # Firestore 저장
    if candidate_id and FIREBASE_ENABLED:
        try:
            db.collection('sourcing_candidates').document(candidate_id).update({
                'shopping_specs':         specs,
                'shopping_specs_total':   shop_total,
                'shopping_specs_updated': fb_fs.SERVER_TIMESTAMP,
            })
        except Exception as e:
            print(f"[Firestore] specs 저장 실패: {e}")

    return jsonify({"specs": specs, "product_match": query, "total": shop_total})


@app.route("/api/ask", methods=["POST"])
def ask_claude():
    data     = request.get_json()
    question = (data.get("question") or "").strip()
    if not question:
        return jsonify({"error": "질문을 입력해주세요"}), 400
    try:
        resp = claude.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1000,
            messages=[{
                "role": "user",
                "content": (
                    "당신은 해외 구매대행/직구 소싱 전문가 어시스턴트입니다. "
                    "아래 질문에 핵심만 간결하게 한국어로 답해주세요. "
                    "불필요한 서론 없이 바로 답변하세요.\n\n"
                    f"질문: {question}"
                ),
            }],
        )
        return jsonify({"answer": resp.content[0].text})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def main():
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 80)))


if __name__ == "__main__":
    main()
